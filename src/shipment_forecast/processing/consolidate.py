"""
Data processing: read supplier Excel files, melt value columns,
consolidate, and write history + output files.
"""
from __future__ import annotations

import re
import shutil
from pathlib import Path
from typing import Iterable

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font
import pandas as pd

from shipment_forecast.paths import SOURCE_DATA_DIR, HISTORY_DIR, OUTPUT_DIR, REPORT_DIR

# ── constants ────────────────────────────────────────────────────────────────

MONTH_ORDER = ["Nov", "Dec", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct"]
MONTH_TO_INT = {m: i for i, m in enumerate(MONTH_ORDER)}  # Nov=0, Oct=11
MONTH_ABBR_RE = re.compile(
    r"\b(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\b", re.IGNORECASE
)

FEATURE_COLS = [
    "SPM", "GTK Suppliers", "Category", "Segment", "Series",
    "Production Year", "Platforms", "Product", "Size", "Color",
    "Location", "ODM", "Supplier Part#", "HP/ODM Part#",
]

VALUE_KEYWORDS = ["Table Price", "Unit Rebate", "Q'ty", "Rebate Amount"]

FY_SHEET_RE = re.compile(r"^FY\d{2}$", re.IGNORECASE)

QUARTER_MONTHS = [
    ("Q1", ["Nov", "Dec", "Jan"]),
    ("Q2", ["Feb", "Mar", "Apr"]),
    ("Q3", ["May", "Jun", "Jul"]),
    ("Q4", ["Aug", "Sep", "Oct"]),
]



# ── helpers ──────────────────────────────────────────────────────────────────


def _normalize_col(col: str) -> str:
    return " ".join(str(col).split())


def _detect_fy_sheets(path: Path) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = [s for s in wb.sheetnames if FY_SHEET_RE.match(s)]
    wb.close()
    return sheets


def common_fy_sheets(excel_paths: list[Path]) -> list[str]:
    """Return FY sheets present in ALL provided Excel files, sorted desc."""
    if not excel_paths:
        return []
    sets = [set(_detect_fy_sheets(p)) for p in excel_paths]
    common = set.intersection(*sets)
    return sorted(common, key=lambda s: int(s[2:]), reverse=True)


def _month_abbr(col_name: str) -> str | None:
    """Extract 3-letter month abbreviation from a column name."""
    m = MONTH_ABBR_RE.search(col_name)
    if m:
        return m.group(0).capitalize()
    return None


def _value_type(col_name: str) -> str | None:
    for kw in VALUE_KEYWORDS:
        if kw.lower() in col_name.lower():
            # normalise key name
            if "table price" in col_name.lower():
                return "Table Price"
            if "unit rebate" in col_name.lower():
                return "Unit Rebate"
            if "q'ty" in col_name.lower() or "qty" in col_name.lower():
                return "Q'ty"
            if "rebate amount" in col_name.lower():
                return "Rebate Amount"
    return None


def _month_index_from_start(month_abbr: str, start_month: str) -> int:
    """Return months since start_month in the fiscal cycle (0 = start)."""
    start_idx = MONTH_TO_INT.get(start_month, 0)
    month_idx = MONTH_TO_INT.get(month_abbr, 0)
    return (month_idx - start_idx) % 12


def _fy_quarter(month_abbr: str) -> str:
    """Map month to Q1-Q4 in fiscal year (Nov=Q1 start)."""
    idx = MONTH_TO_INT[month_abbr]  # Nov=0, Oct=11
    quarter = idx // 3 + 1
    return f"Q{quarter}"


def _year_for_month(fy_str: str, month_abbr: str) -> int:
    fy_year = int(fy_str[2:]) + 2000  # FY26 -> 2026
    if month_abbr in ("Nov", "Dec"):
        return fy_year - 1
    return fy_year


# ── core reader ──────────────────────────────────────────────────────────────


def read_supplier_sheet(path: Path, sheet_name: str, start_month: str) -> pd.DataFrame:
    """
    Read one FY sheet from a supplier Excel file.
    Returns a melted DataFrame with feature + month columns.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        wb.close()
        return pd.DataFrame()

    # Header is row index 1 (0-based), i.e. 2nd row
    raw_header = rows[1]
    # Build header list; forward-fill None (merged cells)
    headers: list[str] = []
    last = ""
    for cell in raw_header:
        val = str(cell).strip() if cell is not None else ""
        if val == "" or val == "None":
            val = last
        else:
            last = val
        headers.append(val)

    data_rows = list(rows[2:])

    # Trim trailing rows that are entirely None — openpyxl often returns thousands
    # of phantom rows from Excel's stored used-range; they explode after ffill × months.
    while data_rows and all(v is None for v in data_rows[-1]):
        data_rows.pop()

    if not data_rows:
        wb.close()
        return pd.DataFrame()

    df_raw = pd.DataFrame(data_rows, columns=headers)

    # ── identify feature and value columns ───────────────────────────────────
    # Match feature columns (flexible: check if header starts-with or equals)
    feature_map: dict[str, str] = {}  # actual_col -> canonical name
    for h in headers:
        h_clean = _normalize_col(h)
        for fc in FEATURE_COLS:
            if h_clean.lower() == fc.lower() or h_clean.lower().startswith(fc.lower()):
                feature_map[h] = fc
                break

    # Value columns: (actual_col) -> (value_type, month_abbr)
    value_map: dict[str, tuple[str, str]] = {}
    for h in headers:
        vt = _value_type(h)
        if vt:
            ma = _month_abbr(h)
            if ma:
                value_map[h] = (vt, ma)

    if not value_map:
        wb.close()
        return pd.DataFrame()

    # Filter to months >= start_month in fiscal cycle
    def keep_month(ma: str) -> bool:
        return _month_index_from_start(ma, start_month) >= 0  # keep all from start

    # months to keep (those at index >=0 from start, up to Oct)
    start_idx = MONTH_TO_INT[start_month]
    months_to_keep = {MONTH_ORDER[(start_idx + i) % 12] for i in range(12 - start_idx)}

    filtered_value_map = {k: v for k, v in value_map.items() if v[1] in months_to_keep}

    if not filtered_value_map:
        wb.close()
        return pd.DataFrame()

    # Build a tidy feature df (forward-fill merged cells in Category/Segment/Series)
    feat_cols_actual = list(feature_map.keys())
    # deduplicate: if same canonical name appears multiple times take first
    seen_canonical: set[str] = set()
    deduped_feat_cols: list[str] = []
    for col in feat_cols_actual:
        canon = feature_map[col]
        if canon not in seen_canonical:
            seen_canonical.add(canon)
            deduped_feat_cols.append(col)

    # Use positional selection to avoid pandas selecting ALL columns with the
    # same name when the Excel header has duplicate entries (merged-cell forward-fill).
    feat_col_positions = [headers.index(c) for c in deduped_feat_cols]
    feat_df = df_raw.iloc[:, feat_col_positions].copy()
    feat_df.columns = [feature_map[c] for c in deduped_feat_cols]

    # Forward-fill merged cell columns
    for col in ("Category", "Segment", "Series"):
        if col in feat_df.columns:
            feat_df[col] = feat_df[col].replace("", pd.NA).ffill()

    # Drop rows where all feature cols are null/empty (phantom trailing rows)
    feat_df = feat_df.dropna(how="all")

    # Secondary guard: drop rows where neither GTK Suppliers nor Supplier Part#
    # has a real value — these are phantom / artifact rows (e.g. "Refresh" buttons).
    # Use GTK Suppliers keyword blocklist FIRST, then require at least one anchor col.
    _ARTIFACTS = {"refresh", "total", "note", "notes", "subtotal",
                  "click", "update", "button", "reset"}

    def _is_artifact(val) -> bool:
        if pd.isna(val):
            return False
        s = str(val).strip()
        return s.lower() in _ARTIFACTS

    if "GTK Suppliers" in feat_df.columns:
        feat_df = feat_df[~feat_df["GTK Suppliers"].apply(_is_artifact)]

    # Require BOTH "Platforms" AND "SPM" to have a real value.
    # Rows missing either are considered header artifacts / summary rows and dropped.
    anchor_must_all = [c for c in ("Platforms", "SPM") if c in feat_df.columns]
    if anchor_must_all:
        has_all = feat_df[anchor_must_all].replace("", pd.NA).notna().all(axis=1)
        feat_df = feat_df[has_all]

    if feat_df.empty:
        wb.close()
        return pd.DataFrame()

    # ── build value columns aligned to surviving feat_df rows ─────────────────
    # IMPORTANT: save original df_raw indices BEFORE any reset_index so that
    # value_df rows are always correctly matched to feat_df rows.
    surviving_idx = feat_df.index  # original positional indices into df_raw
    feat_df = feat_df.reset_index(drop=True)

    value_df = df_raw.iloc[surviving_idx].reset_index(drop=True)

    # ── pivot value cols: one row per (feature combo, month) ─────────────────
    # Group value columns by month, collect {month: {vtype: col}}
    month_vtype_col: dict[str, dict[str, str]] = {}
    for col, (vt, ma) in filtered_value_map.items():
        month_vtype_col.setdefault(ma, {})[vt] = col

    month_dfs: list[pd.DataFrame] = []
    for ma, vtype_cols in month_vtype_col.items():
        m_df = feat_df.copy()
        m_df["Month"] = ma
        for vt in VALUE_KEYWORDS:
            actual_col = vtype_cols.get(vt)
            if actual_col:
                # Use positional indexing to avoid duplicate-name ambiguity
                if actual_col in headers:
                    col_data = value_df.iloc[:, headers.index(actual_col)]
                else:
                    col_data = pd.Series(dtype=float)
                m_df[vt] = pd.to_numeric(col_data, errors="coerce").fillna(0).values
            else:
                m_df[vt] = 0
        month_dfs.append(m_df)

    if not month_dfs:
        wb.close()
        return pd.DataFrame()

    result = pd.concat(month_dfs, ignore_index=True)

    # ── add FY / Quarter / Year cols ─────────────────────────────────────────
    fy = sheet_name.upper()  # e.g. FY26
    fy_int = int(fy[2:])  # 26
    result["FY"] = result["Month"].apply(lambda m: f"{fy} {_fy_quarter(m)}")
    # FY Sort Key: unique integer per FY-Quarter for PowerBI "Sort by column"
    # e.g. FY26 Q1→261, FY26 Q2→262, FY26 Q3→263, FY26 Q4→264
    result["FY Sort Key"] = result["Month"].apply(
        lambda m: fy_int * 10 + int(_fy_quarter(m)[1])
    )
    result["Year"] = result["Month"].apply(lambda m: _year_for_month(fy, m))

    wb.close()
    return result


# ── source data management ────────────────────────────────────────────────────


def clear_and_copy_sources(file_paths: list[Path]) -> list[Path]:
    """Clear source_data dir contents and copy given files into it.

    Avoids shutil.rmtree on the directory itself to prevent WinError 5
    on OneDrive-synced paths where the folder is locked by the sync client.
    """
    SOURCE_DATA_DIR.mkdir(parents=True, exist_ok=True)
    # Delete existing files one by one (don't remove the directory)
    for existing in SOURCE_DATA_DIR.iterdir():
        try:
            if existing.is_file():
                existing.unlink()
            elif existing.is_dir():
                shutil.rmtree(existing, ignore_errors=True)
        except OSError:
            pass  # skip locked files silently

    copied: list[Path] = []
    for src in file_paths:
        dst = SOURCE_DATA_DIR / src.name
        # If duplicate name, prefix with supplier folder name
        if dst.exists():
            dst = SOURCE_DATA_DIR / f"{src.parent.parent.name}_{src.name}"
        shutil.copy2(src, dst)
        copied.append(dst)
    return copied


# ── consolidation ─────────────────────────────────────────────────────────────


def consolidate_monthly(
    source_files: list[Path],
    fy_sheet: str,
    start_month: str,
) -> pd.DataFrame:
    """Read all source files for the given FY sheet and concatenate."""
    dfs: list[pd.DataFrame] = []
    for f in source_files:
        try:
            df = read_supplier_sheet(f, fy_sheet, start_month)
            if not df.empty:
                df["Source File"] = f.name
                dfs.append(df)
        except Exception as exc:
            print(f"[WARN] Failed to read {f.name}: {exc}")
    if not dfs:
        return pd.DataFrame()
    return pd.concat(dfs, ignore_index=True)


def save_to_history(df: pd.DataFrame, fy_sheet: str, start_month: str) -> Path:
    """Save consolidated df to history as 'Rolling Forecast FYXX MM.xlsx'."""
    HISTORY_DIR.mkdir(parents=True, exist_ok=True)
    fy = fy_sheet.upper()  # FY26
    # Convert month abbr to zero-padded number
    month_num = str(MONTH_ORDER.index(start_month) + 1).zfill(2)
    # Fiscal month index: Nov=01 ... Oct=12; convert to calendar-like label
    # Actually the user wants MM = calendar month number of start_month
    import calendar
    cal_abbrs = [calendar.month_abbr[i] for i in range(1, 13)]
    if start_month in cal_abbrs:
        month_num = str(cal_abbrs.index(start_month) + 1).zfill(2)

    # Add Forecast Date label and a numeric sort key for PowerBI.
    # Fiscal order: Nov(11)=1, Dec(12)=2, Jan(1)=3, ..., Oct(10)=12
    forecast_date = f"{fy} {month_num}"
    cal_month = int(month_num)
    fiscal_pos = (cal_month - 11) % 12 + 1   # Nov→1, Dec→2, Jan→3 … Oct→12
    sort_key = int(fy[2:]) * 100 + fiscal_pos  # e.g. FY26 May → 2607

    df = df.copy()
    df["Forecast Date"] = forecast_date
    df["Forecast Date Sort Key"] = sort_key

    fname = f"Rolling Forecast {fy} {month_num}.xlsx"
    out_path = HISTORY_DIR / fname
    sheet_label = f"forecast based on {start_month}"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_label, index=False)
    return out_path


def list_history_files() -> list[Path]:
    """Return Rolling Forecast files in history dir."""
    pattern = re.compile(r"^Rolling Forecast FY\d{2} \d{2}\.xlsx$", re.IGNORECASE)
    return sorted(
        [f for f in HISTORY_DIR.glob("*.xlsx") if pattern.match(f.name)],
        key=lambda f: f.name,
    )


def merge_and_save(
    selected_paths: list[Path],
    output_dir: Path | None = None,
    progress_cb: "Callable[[int], None] | None" = None,
) -> Path:
    """Merge selected history files and write to output_dir/forecast data.xlsx."""
    from typing import Callable  # noqa: F401 – used in type hint string above
    out_dir = output_dir if output_dir else OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)
    dfs: list[pd.DataFrame] = []
    for i, p in enumerate(selected_paths):
        if progress_cb:
            progress_cb(i)
        xl = pd.ExcelFile(p, engine="openpyxl")
        for sheet in xl.sheet_names:
            df = xl.parse(sheet)
            df["_source"] = p.stem
            dfs.append(df)
    if not dfs:
        return out_dir / "forecast data.xlsx"
    merged = pd.concat(dfs, ignore_index=True)
    merged.drop(columns=["_source"], inplace=True, errors="ignore")
    out = out_dir / "forecast data.xlsx"
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        merged.to_excel(writer, sheet_name="Forecast Data", index=False)
    return out


# ── report generation ─────────────────────────────────────────────────────────


def generate_report(history_path: Path, supplier_order: list[str], report_dir: Path | None = None) -> Path:
    """
    Generate Keyboard and Peripheral sheets, each containing two pivot tables
    (Rebate Amount on top, Q'ty below with 3 blank rows gap).
    Rows = Suppliers, Columns = Months + Quarter subtotals (Excel SUM formulas).
    Saved to data/report/forecast pivot FYXX MM.xlsx.
    """
    import calendar as _cal

    df = pd.read_excel(history_path, sheet_name=0)

    # Extract FY year and calendar month number from filename
    m = re.search(r"FY(\d{2})\s+(\d{2})", history_path.stem, re.IGNORECASE)
    fy_year = m.group(1) if m else "??"
    file_month_num = int(m.group(2)) if m else 1
    start_month = _cal.month_abbr[file_month_num]  # e.g. "May"

    # Months from start_month to fiscal year end (Oct), in fiscal order
    start_idx = MONTH_TO_INT.get(start_month, 0)
    all_report_months = [MONTH_ORDER[(start_idx + i) % 12] for i in range(12 - start_idx)]
    months_in_data = set(df["Month"].dropna().unique())
    report_months_set = set(all_report_months)

    # Split data into Peripheral vs Keyboard (case-insensitive on Segment column)
    if "Segment" in df.columns:
        is_peripheral = df["Segment"].fillna("").str.lower().str.strip() == "peripheral"
    else:
        is_peripheral = pd.Series([False] * len(df), index=df.index)

    sheets_data = [
        ("Keyboard",   df[~is_peripheral]),
        ("Peripheral", df[is_peripheral]),
    ]

    # Column plan: interleave month columns with quarter subtotal columns
    col_plan: list[dict] = []
    for q_label, q_months in QUARTER_MONTHS:
        q_present = [mo for mo in q_months if mo in months_in_data and mo in report_months_set]
        if not q_present:
            continue
        for mo in q_present:
            col_plan.append({"type": "month", "abbr": mo, "label": f"{mo}'{fy_year}"})
        col_plan.append({"type": "quarter", "label": q_label, "months": q_present})

    # Output path: "Rolling Forecast FY26 05" → "forecast pivot FY26 05"
    out_name = history_path.stem.replace("Rolling Forecast", "forecast pivot")
    rpt_dir = report_dir if report_dir else REPORT_DIR
    out_path = rpt_dir / f"{out_name}.xlsx"
    rpt_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    _QUARTER_FILL = PatternFill(fill_type="solid", fgColor="C1F0C8")
    _QUARTER_BORDER = Border(
        left=Side(style="thick", color="000000"),
        right=Side(style="thick", color="000000"),
    )
    _DEFAULT_FONT = Font(name="Calibri", size=11)
    _QUARTER_FONT = Font(name="Calibri", size=11, bold=True)
    _CURRENCY_FMT = '$#,##0.00'

    def _write_pivot(ws, df_seg: pd.DataFrame, metric: str, start_row: int) -> int:
        """Write one pivot table starting at start_row. Returns the next free row."""
        is_currency = metric == "Rebate Amount"

        # Build ordered supplier list for this segment (skip suppliers not present)
        sups_in_seg = df_seg["GTK Suppliers"].dropna().unique().tolist()
        seg_lower: dict[str, str] = {s.lower(): s for s in sups_in_seg}
        ordered: list[str] = []
        seen: set[str] = set()
        for s in supplier_order:
            actual = seg_lower.get(s.lower())
            if actual and actual not in seen:
                ordered.append(actual)
                seen.add(actual)
        for s in sups_in_seg:
            if s not in seen:
                ordered.append(s)
                seen.add(s)

        if not ordered:
            return start_row

        pivot = (
            df_seg.groupby(["GTK Suppliers", "Month"])[metric]
            .sum()
            .unstack("Month")
            .reindex(index=ordered)
            .fillna(0)
        )

        # Header row
        hdr_title = ws.cell(row=start_row, column=1, value=metric)
        hdr_title.font = _DEFAULT_FONT
        col_num = 2
        month_to_col: dict[str, int] = {}
        quarter_plan: list[tuple[str, list[int], int]] = []

        for entry in col_plan:
            if entry["type"] == "month":
                mc = ws.cell(row=start_row, column=col_num, value=entry["label"])
                mc.font = _DEFAULT_FONT
                month_to_col[entry["abbr"]] = col_num
                col_num += 1
            else:
                q_month_cols = [month_to_col[mo] for mo in entry["months"]]
                quarter_plan.append((entry["label"], q_month_cols, col_num))
                hdr_cell = ws.cell(row=start_row, column=col_num, value=entry["label"])
                hdr_cell.fill = _QUARTER_FILL
                hdr_cell.border = _QUARTER_BORDER
                hdr_cell.font = _QUARTER_FONT
                col_num += 1

        q_cols = {q_col for _, _, q_col in quarter_plan}

        # Data rows
        for i, supplier in enumerate(ordered):
            row_num = start_row + 1 + i
            ws.cell(row=row_num, column=1, value=supplier).font = _DEFAULT_FONT

            for mo, c in month_to_col.items():
                try:
                    val = float(pivot.loc[supplier, mo])
                except (KeyError, TypeError):
                    val = 0.0
                cell = ws.cell(row=row_num, column=c, value=val)
                cell.font = _DEFAULT_FONT
                if is_currency:
                    cell.number_format = _CURRENCY_FMT

            for q_label, q_month_cols, q_col in quarter_plan:
                if len(q_month_cols) == 1:
                    coord = ws.cell(row=row_num, column=q_month_cols[0]).coordinate
                    formula = f"={coord}"
                else:
                    first = ws.cell(row=row_num, column=q_month_cols[0]).coordinate
                    last = ws.cell(row=row_num, column=q_month_cols[-1]).coordinate
                    formula = f"=SUM({first}:{last})"
                q_cell = ws.cell(row=row_num, column=q_col, value=formula)
                q_cell.fill = _QUARTER_FILL
                q_cell.border = _QUARTER_BORDER
                q_cell.font = _QUARTER_FONT
                if is_currency:
                    q_cell.number_format = _CURRENCY_FMT

        return start_row + 1 + len(ordered)  # next free row

    def _autofit(ws) -> None:
        """Set each column width to fit its widest cell value."""
        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            has_currency = False
            has_number = False
            for cell in col_cells:
                if cell.value is None:
                    continue
                val = str(cell.value)
                if val.startswith("="):
                    # Formula: width depends on whether it's currency or qty
                    if cell.number_format and "$" in cell.number_format:
                        has_currency = True
                    else:
                        has_number = True
                    continue
                if isinstance(cell.value, (int, float)):
                    if cell.number_format and "$" in cell.number_format:
                        has_currency = True
                    else:
                        has_number = True
                    continue
                max_len = max(max_len, len(val))
            # Estimate numeric column width from realistic formatted values
            if has_currency:
                max_len = max(max_len, 13)   # "$1,234,567.89"
            if has_number:
                max_len = max(max_len, 10)   # "12,345,678"
            ws.column_dimensions[col_letter].width = min(max_len + 1, 40)

    for sheet_idx, (sheet_name, df_seg) in enumerate(sheets_data):
        ws = wb.active if sheet_idx == 0 else wb.create_sheet(sheet_name)
        if sheet_idx == 0:
            ws.title = sheet_name

        next_row = _write_pivot(ws, df_seg, "Rebate Amount", start_row=1)
        # 3 blank rows gap
        next_row += 3
        _write_pivot(ws, df_seg, "Q'ty", start_row=next_row)
        _autofit(ws)

    wb.save(out_path)
    return out_path
